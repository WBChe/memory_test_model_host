import binascii
from openpyxl import Workbook
import numpy as np
import matplotlib.pyplot as plt


#compare bin file and output list_compare
def read_bin_file(filename):
    with open(filename, 'rb') as f:
        content = f.read()
    return content

def bin_to_hex(bin_data):
    hex_str = binascii.hexlify(bin_data).decode('utf-8')
    return hex_str

def compare_bin_files(file1, file2):
    bin_data1 = read_bin_file(file1)
    bin_data2 = read_bin_file(file2)

    hex_data1 = bin_to_hex(bin_data1)
    hex_data2 = bin_to_hex(bin_data2)

    # Compare every 8 characters (4 bytes) in the hex strings
    index = 0
    differences = []

    while index < len(hex_data1) and index < len(hex_data2):
        if hex_data1[index:index+8] != hex_data2[index:index+8]:
            # differences.append(hex_data1[index:index+8])
            differences.append(hex_data2[index:index+8])
        index += 8

    return differences

#reverse list
def reverse_elements_in_list(lst):
    return [element[::-1] for element in lst]

def swap_characters(s):
    if len(s) < 2:
        return s
    else:
        swapped = s[1] + s[0] + swap_characters(s[2:])
        return swapped
    
def swap_characters_in_list(lst):
    return [swap_characters(element) for element in lst]

#remove 2head 2tile
def remove_first_last_two(input_str):
    # 去掉前两个字符和后两个字符
    return input_str[2:-2]

def transform_list(lst):
    transformed_list = [remove_first_last_two(item) for item in lst]
    return transformed_list


def get_fault_addr(file_result, file_check):
    differences = compare_bin_files(file_result, file_check)
    # 列表中每个元素倒序排列
    reversed_list = reverse_elements_in_list(differences)
    # 列表中每个元素的字符两两交换
    new_compare_data = swap_characters_in_list(reversed_list)
    # 去掉前两个字符和后两个字符
    fault_addr = transform_list(new_compare_data)
    # string 2 int hex
    fault_addr_hex = [int(element, 16) for element in fault_addr]   

    return fault_addr_hex


# prepare for get range
def split_range_into_parts(start, end, parts):
    # Calculate the step size
    step = (end - start + 1) / parts  # Adding 1 to end makes end inclusive
    
    # Generate the split points
    split_points = []
    current = start
    for i in range(parts):
        start_val = int(current)
        current += step
        end_val = int(current) - 1  # end_val is inclusive
        split_points.append((start_val, end_val, i + 1))  # i + 1 to make index start from 1
    
    return split_points

def get_range_block(start_hex, end_hex, num_parts): 

    # Define the range in hexadecimal
    # start_hex = 0x0000
    # end_hex = 0x7fff

    # # Number of parts to split into
    # num_parts = 1024

    # Convert hexadecimal to decimal for calculations
    start_decimal = int(start_hex)
    end_decimal = int(end_hex)

    # Split the range into parts with start, end, and index
    split_parts = split_range_into_parts(start_decimal, end_decimal, num_parts)

    # Convert the results to the desired output format
    desired_output = [(start, end, index) for (start, end, index) in split_parts]

    return desired_output

# get range
def get_label(element):
    # 定义范围和对应的标号
    ranges = get_range_block(0x0000, 0x7fff, 1024)  #----------------------------------------------------------you can change block divide
    print(ranges)
    
    # 遍历范围，查找匹配的标号
    for start, end, label in ranges:
        if start <= element <= end: # detect range
            return label
    
    # 如果没有找到匹配的范围，返回 None 或者其他你认为合适的值
    return None

# 列表内自己remove duplicate
def remove_duplicates(lst):
    seen = set()
    result = []
    for item in lst:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result

# write excel
def write_to_excel(data, filename):
    # 创建一个新的 Workbook 对象
    wb = Workbook()
    # 激活第一个工作表
    ws = wb.active
    
    # 遍历二维列表，并写入到工作表中
    for row in data:
        ws.append(row)
    
    # 保存工作簿到指定的文件名
    wb.save(filename)
    print(f"数据已写入到 {filename} 文件中。\n")

# write excel order version
def write_to_excel_order(data, filename):
    # 创建一个新的工作簿
    wb = Workbook()

    # 激活（选择）工作表
    ws = wb.active

    for j in range(0, 14):
    # 写入数据
        for i, value in enumerate(data[j]):
            # 写入第一个元素到第一列
            if i == 0:
                ws.cell(row=j+1, column=1, value=value)
            else:
                # 计算列号，第二个元素放在第(value + 1)列
                column_index = value + 1
                ws.cell(row=j+1, column=column_index, value=value)

    # 保存工作簿
    wb.save(filename)

    print(f"数据已成功写入 Excel 文件 '{filename}'")


def compare_and_filter_lists(lists):
    # 从第三个元素开始取出来与其他列表对比
    for i in range(len(lists)):
        current_list = lists[i]
        for j in range(2, len(current_list)):  
            current_element = current_list[j]
            for k in range(len(lists)):
                if k == i:
                    continue  # 不与自己比较
                other_list = lists[k]
                if current_element in other_list:
                    # 找到了相同的元素，比较第二个元素
                    if current_list[1] > other_list[1]:
                        # 如果当前列表第二个元素值大，continue
                        continue
                    else:
                        # 如果其他列表第二个元素值大，则删除其他列表中的元素
                        other_list.remove(current_element)

# make new pattern
def update_list(main_list, update_list):
    if len(update_list) < 2:
        print("Update list must have at least two elements (value and index pairs).")
        return
    
    value_to_change = update_list[0]
    for index in update_list[1:]:
        if 0 <= index < len(main_list):
            main_list[index-1] = value_to_change

def make_new_pattern(new_algorithm_block):
    compare_and_filter_lists(new_algorithm_block)
    print(new_algorithm_block)

    #delete algorithm time
    for sublist in new_algorithm_block:
        del sublist[1]
    print(new_algorithm_block)

    # make new pattern for every block
    new_pattern = ['MSCAN'] * 1024

    for i in range(0, len(new_algorithm_block)):
        update_list(new_pattern, new_algorithm_block[i])

    return new_pattern

# visualize new pattern 
def visualize_pattern(new_pattern, char_to_int, char_to_abc):
    # 用于后面矩阵填充字母
    new_pattern_mapping = [char_to_abc[c] for c in new_pattern]
    # 将字符串列表映射为整数数组，用于表示颜色或灰度值
    int_array = np.array([char_to_int[c] for c in new_pattern]).reshape(32, 32)

    # 创建一个图形窗口并绘制图像
    plt.figure(figsize=(6, 6))  # 设置图形大小为 6x6 英寸
    plt.imshow(int_array, cmap='tab20', interpolation='nearest')    # tab20 调色板20种颜色


    # 添加每个像素的数据类型标签
    # for i in range(int_array.shape[0]):
    #     for j in range(int_array.shape[1]):
    #         plt.text(j, i, new_pattern[i * int_array.shape[1] + j], ha='center', va='center', color='black')

    # 添加每个像素的数据类型标签 A B C ...
    for i in range(int_array.shape[0]):
        for j in range(int_array.shape[1]):
            plt.text(j, i, new_pattern_mapping[i * int_array.shape[1] + j], ha='center', va='center', color='black') #new_pattern #new_pattern[i * int_array.shape[1] + j]

    # 创建自定义颜色条来表示数据类型
    legend_elements = [
        plt.Line2D([0], [0], marker='o', color='w', label='Type A :   MSCAN', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type B :   Checkerboard', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type C :   MATS', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type D :   MATS+', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type E :   MATS++', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type F :   March A', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type G :   March B', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type H :   March C', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type I :   March C-', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type J :   March C+', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type K :   March U', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type L :   March LR', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type M :   March SR', markersize=10),
        plt.Line2D([0], [0], marker='o', color='w', label='Type N :   March SS', markersize=10),
    ]
    plt.legend(handles=legend_elements, loc='center left', bbox_to_anchor=(1, 0.5))  # 将图例放在图像右侧

    plt.title('32x32 new recombination pattern')
    plt.axis('off')  # 关闭坐标轴
    plt.show()


def get_mutifile_fault(algorithm_name, test_num):
    one_fault_addr_array = [[] for _ in range(test_num)] # creat 2d list
    one_algorithm_fault_block = [] # creat 2d list
    for i in range(0, test_num): # algorithm_num+1
        print('now algorithm is : ' + algorithm_name )   
        algorithm_result_bin = './result/' + algorithm_name +'_result.bin'
        algorithm_check_bin = './check/' + algorithm_name + '_check_' + str(i) + '.bin'

        one_fault_addr_array[i] = get_fault_addr(algorithm_result_bin, algorithm_check_bin)
        print(one_fault_addr_array[i])

        # get block list 
        for elem in one_fault_addr_array[i]:
            label = get_label(elem)     # block decide
            if label is not None:
                one_algorithm_fault_block.append(label)
    
    return one_algorithm_fault_block

def get_muti_algorithm_fault(algorithm_name_list, test_num):
    muti_algorithm_fault_block = []
    for name in algorithm_name_list:
        one_algorithm_fault_block = get_mutifile_fault(name, test_num)
        muti_algorithm_fault_block.append(one_algorithm_fault_block)
    return muti_algorithm_fault_block












#---------------------------------------------------------------------main----------------------------------------------------------------------------#

if __name__ == "__main__":

    # # define file source
    # algorithm_result_bin    = []
    # algorithm_check_bin     = []

    # algorithm_result_bin.append('./result/mscan_result.bin')
    # algorithm_result_bin.append('./result/checkerboard_result.bin')

    # algorithm_check_bin.append('./check/mscan_check_0.bin')
    # algorithm_check_bin.append('./check/checkerboard_check_0.bin')

    # # init parameter and list
    # algorithm_num = 14
    # fault_addr_array = [[] for _ in range(algorithm_num)] # creat 2d list
    # algorithm_fault_block = [[] for _ in range(algorithm_num)] # creat 2d list
    # new_algorithm_fault_block = [[] for _ in range(algorithm_num)] # creat 2d list
    # print('init fault_addr_array :')
    # print(fault_addr_array)
    # print('\n')
    
    # # get fault block list 复用于一个算法得到多个测试结果
    # for i in range(0, len(algorithm_result_bin)): # algorithm_num+1
    #     print('now algorithm is : ' + str(i))   
    #     print(algorithm_result_bin[i])  
    #     print(algorithm_check_bin[i]) 
    #     fault_addr_array[i] = get_fault_addr(algorithm_result_bin[i], algorithm_check_bin[i])
    #     print(fault_addr_array[i])

    #     # get block list 
    #     for elem in fault_addr_array[i]:
    #         label = get_label(elem)     # block decide
    #         if label is not None:
    #             algorithm_fault_block[i].append(label)

    # 
    algorithm_name_list_all = ['MSCAN', 'Checkerboard', 'MATS', 'MATS+', 'MATS++', 'March_A', 'March_B', 'March_C', 'March_C-', 'March_C+', 'March_U', 'March_LR', 'March_SR', 'March_SS']
    algorithm_name_list = ['MSCAN', 'Checkerboard']
    algorithm_time = [4, 4, 4, 5, 6, 15, 17, 11, 10, 14, 13, 14, 14, 22]    # algorithm time N

    test_num = 2
    muti_algorithm_fault_block = get_muti_algorithm_fault(algorithm_name_list, test_num)    # 遍历check与result对比得到多个test的结果

    print('this is algorithm_fault_block:')
    print(muti_algorithm_fault_block)    #-----------------------------------------------------想直接通过函数得到多个测试的二维列表，下面继续程序继续使用，上面修改


    
    # remove duplicates 
    new_algorithm_fault_block = [[] for _ in range(14)] # creat 2d list

    for i in range(0, len(muti_algorithm_fault_block)):
        new_algorithm_fault_block[i] = remove_duplicates(muti_algorithm_fault_block[i])
    
    print(new_algorithm_fault_block)

    # add algorithm_name, algorithm_time
    for i in range(0, 14):   # len(algorithm_result_bin) 14
        new_algorithm_fault_block[i].sort()
        new_algorithm_fault_block[i].insert(0, algorithm_time[i])
        new_algorithm_fault_block[i].insert(0, algorithm_name_list_all[i])

    print(new_algorithm_fault_block)

    # make excel
    write_to_excel(new_algorithm_fault_block, "test_data.xlsx")
    
    # make new pattern
    new_pattern = make_new_pattern(new_algorithm_fault_block)

    print('this is new memory test pattern : ')
    print(new_pattern)  #------------------------------------------------------------------this is new pattern


    # visualize new pattern
    # 创建一个映射字典将字符映射到整数
    char_to_int = { 'MSCAN'        : 1, 
                    'Checkerboard' : 2,
                    'MATS'         : 3,
                    'MATS+'        : 4,
                    'MATS++'       : 5,
                    'March_A'      : 6,
                    'March_B'      : 7,
                    'March_C'      : 8,
                    'March_C-'     : 9,
                    'March_C+'     : 10,
                    'March_U'      : 11,
                    'March_LR'     : 12,
                    'March_SR'     : 13,
                    'March_SS'     : 14}  
   
    # 创建一个映射字典将字符映射到字母
    char_to_abc = { 'MSCAN'        : 'A', 
                    'Checkerboard' : 'B',
                    'MATS'         : 'C',
                    'MATS+'        : 'D',
                    'MATS++'       : 'E',
                    'March_A'      : 'F',
                    'March_B'      : 'G',
                    'March_C'      : 'H',
                    'March_C-'     : 'I',
                    'March_C+'     : 'J',
                    'March_U'      : 'K',
                    'March_LR'     : 'L',
                    'March_SR'     : 'M',
                    'March_SS'     : 'N'}  
    # visualize function
    visualize_pattern(new_pattern, char_to_int, char_to_abc)


    print('get make new pattern done')


    









